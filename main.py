import sys
import os
import random
import json
import threading
import shutil
import zipfile
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QTableWidget, QTableWidgetItem, QPushButton, QLineEdit, 
                             QComboBox, QSpinBox, QFileDialog, QMessageBox, QLabel,
                             QCheckBox, QTextEdit, QGroupBox, QHeaderView, 
                             QSystemTrayIcon, QMenu, QAction, QDialog,
                             QDialogButtonBox, QStatusBar, QScrollArea,
                             QGridLayout, QFrame, QProgressBar, QProgressDialog,
                             QListWidget, QListWidgetItem)
from PyQt5.QtGui import QIcon, QMovie, QPixmap
import subprocess
import tempfile
from openpyxl import load_workbook

# 路径处理：config.json放在exe所在目录
if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
    # PyInstaller内置资源目录
    BUNDLE_DIR = getattr(sys, '_MEIPASS', APP_DIR)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
    BUNDLE_DIR = APP_DIR

CONFIG_FILE = os.path.join(APP_DIR, 'config.json')
# 优先从exe目录找Excel，其次从内置资源目录找
DEFAULT_EXCEL_PATH = os.path.join(APP_DIR, 'BUFF动画职业名对照表.xlsx')
if not os.path.exists(DEFAULT_EXCEL_PATH):
    DEFAULT_EXCEL_PATH = os.path.join(BUNDLE_DIR, 'BUFF动画职业名对照表.xlsx')


class PresetNameDialog(QDialog):
    """自定义预设命名对话框，替代QInputDialog避免打包后阻塞"""
    def __init__(self, parent=None, title='保存预设', label='输入预设名称:', default=''):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(label))
        self.input = QLineEdit(default)
        layout.addWidget(self.input)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)
    
    def get_name(self):
        return self.input.text().strip()


def find_ffmpeg():
    """查找ffmpeg：优先用imageio-ffmpeg内置的，再找程序目录和PATH"""
    # 1. imageio-ffmpeg 内置（最可靠，打包后也有效）
    try:
        import imageio_ffmpeg
        return imageio_ffmpeg.get_ffmpeg_exe()
    except Exception:
        pass
    # 2. 程序目录
    local_ffmpeg = os.path.join(APP_DIR, 'ffmpeg.exe')
    if os.path.isfile(local_ffmpeg):
        return local_ffmpeg
    # 3. PATH
    import shutil as _sh
    return _sh.which('ffmpeg')


class Bk2PreviewDialog(QDialog):
    """Bk2动画预览对话框：用ffmpeg转gif后用QMovie播放"""
    def __init__(self, bk2_path, parent=None):
        super().__init__(parent)
        name = os.path.basename(bk2_path)
        self.setWindowTitle(f'Buff预览 - {name}')
        self.setMinimumSize(100, 100)
        self.bk2_path = bk2_path
        self.gif_path = None
        self.movie = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # 预览区域
        self.preview_label = QLabel('转换中...')
        self.preview_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(self.preview_label)

        # 开始转换
        QtCore.QTimer.singleShot(100, self.convert_to_gif)

    def convert_to_gif(self):
        ffmpeg = find_ffmpeg()
        if not ffmpeg:
            self.preview_label.setText(
                '未找到 ffmpeg.exe！\n\n'
                '请下载 ffmpeg.exe 放到程序同目录下，\n'
                '或点击下方"用系统默认程序打开"。\n\n'
                'ffmpeg下载: https://ffmpeg.org/download.html')
            return

        # 临时gif
        self.gif_path = os.path.join(tempfile.gettempdir(), f'bk2_preview_{os.getpid()}.gif')
        if os.path.exists(self.gif_path):
            os.remove(self.gif_path)

        # 调用ffmpeg转换: 10fps, 保持原始尺寸, 最多5秒
        cmd = [ffmpeg, '-y', '-i', self.bk2_path,
               '-vf', 'fps=10',
               '-t', '5', '-loop', '0', self.gif_path]

        try:
            result = subprocess.run(cmd, capture_output=True, timeout=30,
                                   creationflags=subprocess.CREATE_NO_WINDOW)
            if result.returncode != 0 or not os.path.exists(self.gif_path):
                err = result.stderr.decode('utf-8', errors='ignore')[-500:]
                self.preview_label.setText(f'转换失败!\n\nffmpeg错误:\n{err}')
                return
        except Exception as e:
            self.preview_label.setText(f'转换失败!\n\n错误: {e}')
            return

        # 用QMovie播放
        self.movie = QMovie(self.gif_path)
        if not self.movie.isValid():
            self.preview_label.setText('无法加载gif，请用系统默认程序打开。')
            return
        self.movie.setCacheMode(QMovie.CacheAll)
        self.movie.frameChanged.connect(self.on_frame_changed)
        self.movie.start()
        self.preview_label.setMovie(self.movie)
        self.preview_label.setText('')

    def on_frame_changed(self, frame_number):
        """帧变化时调整窗口大小"""
        if self.movie and self.movie.state() == QMovie.Running:
            frame_pixmap = self.movie.currentPixmap()
            if not frame_pixmap.isNull():
                actual_size = frame_pixmap.size()
                if actual_size.width() > 0 and actual_size.height() > 0:
                    self.preview_label.setFixedSize(actual_size)
                    self.adjustSize()
                    self.movie.frameChanged.disconnect(self.on_frame_changed)

    def open_external(self):
        try:
            os.startfile(self.bk2_path)
        except Exception as e:
            QMessageBox.warning(self, '错误', f'无法打开文件:\n{e}')

    def closeEvent(self, event):
        if self.movie:
            self.movie.stop()
        if self.gif_path and os.path.exists(self.gif_path):
            try:
                os.remove(self.gif_path)
            except Exception:
                pass
        event.accept()


class Bk2GalleryDialog(QDialog):
    """备用库画廊：展示所有bk2缩略图，可分配给职业"""
    def __init__(self, backup_path, thumb_dir, rows_info, parent=None):
        super().__init__(parent)
        self.setWindowTitle('备用动画库 - 框选/Ctrl+A/Shift选择')
        self.setMinimumSize(740, 740)
        self.backup_path = backup_path
        self.thumb_dir = thumb_dir
        self.rows_info = rows_info
        self.main_window = parent
        self.preview_mode = 'standard'

        self.STANDARD_SIZE = (180, 140)
        self.LARGE_SIZE = (360, 280)

        layout = QVBoxLayout(self)

        top_bar = QHBoxLayout()
        top_label = QLabel('支持框选、Ctrl+A全选、Ctrl+点击多选、Shift+点击范围选')
        top_label.setStyleSheet('color: #555; padding: 4px;')
        top_bar.addWidget(top_label)
        top_bar.addStretch()

        self.mode_btn = QPushButton('大图预览')
        self.mode_btn.clicked.connect(self.toggle_mode)
        top_bar.addWidget(self.mode_btn)

        self.select_all_btn = QPushButton('全选')
        self.select_all_btn.clicked.connect(self.select_all)
        top_bar.addWidget(self.select_all_btn)
        self.deselect_all_btn = QPushButton('取消全选')
        self.deselect_all_btn.clicked.connect(self.deselect_all)
        top_bar.addWidget(self.deselect_all_btn)
        layout.addLayout(top_bar)

        # 使用QListWidget实现任务管理器式选择
        self.list_widget = QListWidget()
        self.list_widget.setViewMode(QListWidget.IconMode)
        self.list_widget.setSelectionMode(QListWidget.ExtendedSelection)
        self.list_widget.setResizeMode(QListWidget.Adjust)
        self.list_widget.setSpacing(8)
        self.list_widget.setDragEnabled(False)
        self.list_widget.setDragDropMode(QListWidget.NoDragDrop)
        self.list_widget.itemSelectionChanged.connect(self.on_selection_changed)
        self.list_widget.itemDoubleClicked.connect(self.on_item_double_clicked)
        layout.addWidget(self.list_widget)

        # 底部分配栏
        bottom = QHBoxLayout()
        self.selected_label = QLabel('未选择文件')
        bottom.addWidget(self.selected_label)
        bottom.addStretch()

        bottom.addWidget(QLabel('动画分组:'))
        self.vocation_combo = QComboBox()
        groups = {}
        for info in rows_info:
            keyword = info.get('keyword', '')
            if keyword and keyword not in groups:
                vocations = [i['vocation'] for i in rows_info if i.get('keyword') == keyword]
                groups[keyword] = vocations
        for keyword, vocations in groups.items():
            label = f"{keyword} ({', '.join(vocations)})"
            self.vocation_combo.addItem(label, keyword)
        bottom.addWidget(self.vocation_combo)

        self.assign_btn = QPushButton('分配（复制到动画路径）')
        self.assign_btn.clicked.connect(self.assign_to_vocation)
        self.assign_btn.setEnabled(False)
        bottom.addWidget(self.assign_btn)

        self.preview_btn = QPushButton('预览动画')
        self.preview_btn.clicked.connect(self.preview_selected)
        self.preview_btn.setEnabled(False)
        bottom.addWidget(self.preview_btn)

        close_btn = QPushButton('关闭')
        close_btn.clicked.connect(self.close)
        bottom.addWidget(close_btn)

        layout.addLayout(bottom)

        self.load_thumbnails()

    def toggle_mode(self):
        """切换预览模式"""
        saved_selection = [item.data(QtCore.Qt.UserRole) for item in self.list_widget.selectedItems()]
        if self.preview_mode == 'standard':
            self.preview_mode = 'large'
            self.mode_btn.setText('标准预览')
        else:
            self.preview_mode = 'standard'
            self.mode_btn.setText('大图预览')
        self.load_thumbnails()
        for item in self.list_widget.findItems("", QtCore.Qt.MatchContains):
            bk2 = item.data(QtCore.Qt.UserRole)
            if bk2 in saved_selection:
                item.setSelected(True)

    def load_thumbnails(self):
        """加载所有bk2缩略图"""
        self.list_widget.clear()
        bk2_files = sorted([f for f in os.listdir(self.backup_path)
                           if f.lower().endswith('.bk2')])
        if not bk2_files:
            self.selected_label.setText('备用路径中没有.bk2文件')
            return

        size = self.LARGE_SIZE if self.preview_mode == 'large' else self.STANDARD_SIZE
        img_size = (size[0] - 10, int(size[1] * 0.68))

        self.list_widget.setIconSize(QtCore.QSize(img_size[0], img_size[1]))
        self.list_widget.setGridSize(QtCore.QSize(size[0], size[1]))

        for bk2 in bk2_files:
            thumb_name = os.path.splitext(bk2)[0] + '.png'
            thumb_path = os.path.join(self.thumb_dir, thumb_name)

            if os.path.exists(thumb_path):
                pix = QPixmap(thumb_path)
                if not pix.isNull():
                    icon = QIcon(pix.scaled(img_size[0], img_size[1],
                                           QtCore.Qt.KeepAspectRatio,
                                           QtCore.Qt.SmoothTransformation))
                else:
                    icon = QIcon()
            else:
                icon = QIcon()

            item = QListWidgetItem(icon, bk2)
            item.setData(QtCore.Qt.UserRole, bk2)
            item.setTextAlignment(QtCore.Qt.AlignHCenter)
            item.setToolTip(bk2)
            self.list_widget.addItem(item)

    def on_selection_changed(self):
        """选择变化时更新UI"""
        selected_items = self.list_widget.selectedItems()
        count = len(selected_items)
        if count == 0:
            self.selected_label.setText('未选择文件')
            self.selected_label.setStyleSheet('color: #555; padding: 4px;')
            self.assign_btn.setEnabled(False)
            self.preview_btn.setEnabled(False)
        else:
            self.selected_label.setText(f'已选择 {count} 个文件')
            self.selected_label.setStyleSheet('color: #2196F3; font-weight: bold; padding: 4px;')
            self.assign_btn.setEnabled(True)
            self.preview_btn.setEnabled(count == 1)

    def on_item_double_clicked(self, item):
        """双击预览动画"""
        bk2_name = item.data(QtCore.Qt.UserRole)
        bk2_path = os.path.join(self.backup_path, bk2_name)
        dlg = Bk2PreviewDialog(bk2_path, self)
        dlg.exec_()

    def select_all(self):
        """全选"""
        self.list_widget.selectAll()

    def deselect_all(self):
        """取消全选"""
        self.list_widget.clearSelection()

    def preview_selected(self):
        """预览选中的bk2（仅单选时可用）"""
        selected_items = self.list_widget.selectedItems()
        if len(selected_items) != 1:
            return
        bk2_name = selected_items[0].data(QtCore.Qt.UserRole)
        bk2_path = os.path.join(self.backup_path, bk2_name)
        dlg = Bk2PreviewDialog(bk2_path, self)
        dlg.exec_()

    def assign_to_vocation(self):
        """将选中的多个bk2复制到动画路径，以动画分组前缀命名"""
        selected_items = self.list_widget.selectedItems()
        if not selected_items:
            return

        keyword = self.vocation_combo.currentData()
        if not keyword:
            QMessageBox.warning(self, '警告', '请先在主页添加动画分组!')
            return

        target_filepath = self.main_window.global_filepath

        if not target_filepath or not os.path.isdir(target_filepath):
            QMessageBox.warning(self, '警告', '主动画路径未设置!')
            return

        current_text = self.vocation_combo.currentText()

        # 当前分组包含的所有职业
        group_vocations = [i['vocation'] for i in self.rows_info if i.get('keyword') == keyword]

        # 弹窗1：是否删除原版Buff动画
        delete_original = False
        if group_vocations:
            orig_files = []
            for voc in group_vocations:
                orig = self.main_window.vocation_mapping.get_target_filename(voc)
                if orig and orig not in orig_files:
                    orig_files.append(orig)
            if orig_files:
                ret = QMessageBox.question(
                    self, '删除原版Buff动画',
                    f'是否删除当前分组的原版Buff动画？\n\n'
                    f'分组: {current_text}\n'
                    f'将删除以下原版文件:\n'
                    f'{chr(10).join("- " + f for f in orig_files)}\n\n'
                    f'选择"是"将删除这些原版动画文件；\n'
                    f'选择"否"则保留原版文件。',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                delete_original = (ret == QMessageBox.Yes)

        # 弹窗2：确认分配
        assign_msg = (
            f'即将将 {len(selected_items)} 个文件复制到:\n\n'
            f'{target_filepath}\n\n'
            f'动画分组: {keyword}\n'
            f'关联职业: {current_text}\n'
            f'文件名格式: {keyword}随机数.bk2\n\n'
        )
        if delete_original:
            assign_msg += '[注意] 将先删除该分组所选职业的原版Buff动画文件\n\n'
        assign_msg += '是否继续?'

        confirm = QMessageBox.question(
            self, '确认分配',
            assign_msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        # 复制前：删除原版Buff动画文件
        deleted = []
        del_fail = []
        if delete_original:
            def _remove_original(orig):
                p = os.path.join(target_filepath, orig)
                if os.path.exists(p):
                    try:
                        os.remove(p)
                        return True
                    except Exception as e:
                        del_fail.append(f'{orig}: {e}')
                return False
            for voc in group_vocations:
                orig = self.main_window.vocation_mapping.get_target_filename(voc)
                if orig and _remove_original(orig):
                    deleted.append(orig)

        # 批量复制
        success = 0
        failed = []
        for item in selected_items:
            bk2_name = item.data(QtCore.Qt.UserRole)
            src = os.path.join(self.backup_path, bk2_name)
            if keyword:
                new_name = f"{keyword}{random.randint(1000, 9999)}.bk2"
            else:
                new_name = bk2_name
            dst = os.path.join(target_filepath, new_name)
            try:
                shutil.copy2(src, dst)
                success += 1
            except Exception as e:
                failed.append(f'{bk2_name}: {e}')

        msg = f'复制完成! 成功: {success}/{len(selected_items)}'
        if delete_original:
            msg += f'\n\n已删除原版文件: {len(deleted)} 个'
            if del_fail:
                msg += f'\n删除失败:\n' + '\n'.join(del_fail[:5])
        if failed:
            msg += f'\n\n复制失败:\n' + '\n'.join(failed[:5])
        QMessageBox.information(self, '完成', msg)

        status = f'已分配 {success} 个文件到动画路径'
        if delete_original:
            status += f'，删除原版 {len(deleted)} 个'
        self.main_window.show_status(status)


class GlobalFileManager:
    """全局文件管理器，解决多职业同关键词冲突"""
    _instance = None
    _lock = threading.Lock()
    
    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        self._initialized = True
        self.keyword_to_vocations = {}
        self.vocation_configs = {}
        self._op_lock = threading.Lock()
    
    def register(self, vocation, config):
        with self._op_lock:
            self.vocation_configs[vocation] = config
            keyword = config.get('keyword', '')
            if keyword not in self.keyword_to_vocations:
                self.keyword_to_vocations[keyword] = []
            if vocation not in self.keyword_to_vocations[keyword]:
                self.keyword_to_vocations[keyword].append(vocation)
    
    def unregister(self, vocation):
        with self._op_lock:
            if vocation in self.vocation_configs:
                config = self.vocation_configs[vocation]
                keyword = config.get('keyword', '')
                if keyword in self.keyword_to_vocations:
                    if vocation in self.keyword_to_vocations[keyword]:
                        self.keyword_to_vocations[keyword].remove(vocation)
                del self.vocation_configs[vocation]
    
    def is_keyword_shared(self, keyword):
        with self._op_lock:
            return len(self.keyword_to_vocations.get(keyword, [])) > 1
    
    def get_occupied_targets(self, keyword, exclude_vocation=''):
        with self._op_lock:
            occupied = set()
            for voc, cfg in self.vocation_configs.items():
                if voc != exclude_vocation and cfg.get('keyword', '') == keyword:
                    t = cfg.get('target_filename', '')
                    if t:
                        occupied.add(t)
            return occupied


class BuffSwitcher(QtCore.QObject):
    def __init__(self, config, file_manager):
        super().__init__()
        self.config = config
        self.file_manager = file_manager
        self.running = False
        self.timer = None
        
    def switch(self):
        try:
            filepath = self.config.get('filepath', '')
            target_filename = self.config.get('target_filename', '')
            keyword = self.config.get('keyword', '')
            vocation = self.config.get('vocation', '')
            
            if not os.path.isdir(filepath) or not target_filename or not keyword:
                return False
            
            target_path = os.path.join(filepath, target_filename)
            backup_files = []
            
            # 如果目标文件已存在，先重命名为备用（解除占用）
            if os.path.exists(target_path):
                new_name = self.generate_unique_name(filepath, keyword) + '.bk2'
                new_path = os.path.join(filepath, new_name)
                os.rename(target_path, new_path)
                backup_files.append(new_path)
            
            # 收集所有含关键词的.bk2文件
            occupied = self.file_manager.get_occupied_targets(keyword, vocation)
            for f in os.listdir(filepath):
                if (keyword.lower() in f.lower() 
                    and f.lower().endswith('.bk2') 
                    and f != target_filename
                    and f not in occupied):
                    backup_files.append(os.path.join(filepath, f))
            
            if backup_files:
                random_file = random.choice(backup_files)
                os.rename(random_file, target_path)
                return True
            else:
                print(f"[{vocation}] 警告: 没有找到可用文件")
                return False
        except Exception as e:
            print(f"[{vocation}] 切换失败: {e}")
            return False
    
    def generate_unique_name(self, filepath, keyword):
        existing = set(os.listdir(filepath))
        for _ in range(100):
            name = f"{keyword}{random.randint(1000, 9999)}"
            if name not in existing:
                return name
        return f"{keyword}{random.randint(10000, 99999)}"
    
    def start(self):
        if self.running:
            return
        self.running = True
        self.switch()
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.switch)
        self.timer.start(self.config.get('interval', 30) * 1000)
    
    def stop(self):
        self.running = False
        if self.timer:
            self.timer.stop()
            self.timer = None


class VocationMapping:
    def __init__(self):
        self.mappings = {}
    
    def load_from_excel(self, filepath):
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not header_row:
                print("导入Excel失败: 表头为空")
                return False
            
            vocation_col_idx = -1
            filename_col_idx = -1
            for idx, cell in enumerate(header_row):
                if cell and '职业' in str(cell):
                    vocation_col_idx = idx
                elif cell and '文件' in str(cell):
                    filename_col_idx = idx
            
            if vocation_col_idx == -1 or filename_col_idx == -1:
                print(f"导入Excel失败: 未找到职业名列或文件名列，表头: {header_row}")
                return False
            
            self.mappings = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row:
                    if len(row) > max(vocation_col_idx, filename_col_idx):
                        vocation_cell = row[vocation_col_idx]
                        filename_cell = row[filename_col_idx]
                        
                        if vocation_cell and filename_cell:
                            vocation = str(vocation_cell).strip()
                            target_filename = str(filename_cell).strip()
                            
                            if vocation and target_filename:
                                if '【' in vocation or '】' in vocation:
                                    continue
                                self.mappings[vocation] = target_filename
            
            return True
        except Exception as e:
            print(f"导入Excel失败: {e}")
            return False
    
    def get_vocations(self):
        return list(self.mappings.keys())
    
    def get_target_filename(self, vocation):
        filename = self.mappings.get(vocation, '')
        if filename and not filename.lower().endswith('.bk2'):
            return filename + '.bk2'
        return filename


class PresetManager:
    def __init__(self):
        self.presets = {}
        self.last_preset = ''
        self.last_error = ''
        self.global_settings = {}
        self.load_presets()
    
    def load_presets(self):
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.presets = data.get('presets', {})
                    self.last_preset = data.get('last_preset', '')
                    self.global_settings = data.get('global_settings', {})
        except Exception as e:
            self.last_error = str(e)
            self.presets = {}
            self.global_settings = {}
    
    def save_presets(self):
        try:
            # 先序列化到内存，确保数据可序列化
            data = {
                'presets': self.presets,
                'last_preset': self.last_preset,
                'global_settings': self.global_settings
            }
            json_str = json.dumps(data, ensure_ascii=False, indent=2)
            
            # 写入临时文件再替换，避免写入失败导致配置损坏
            tmp_file = CONFIG_FILE + '.tmp'
            with open(tmp_file, 'w', encoding='utf-8') as f:
                f.write(json_str)
            
            # 替换原文件
            if os.path.exists(CONFIG_FILE):
                os.replace(tmp_file, CONFIG_FILE)
            else:
                os.rename(tmp_file, CONFIG_FILE)
            
            self.last_error = ''
            return True
        except Exception as e:
            self.last_error = str(e)
            return False
    
    def set_last_preset(self, name):
        self.last_preset = name
        self.save_presets()
    
    def save_global_settings(self, settings, merge=False):
        """单独保存全局设置，不依赖预设
        merge=True时合并到现有设置，避免覆盖丢失其他配置
        """
        if merge:
            if not self.global_settings:
                self.global_settings = {}
            self.global_settings.update(settings)
        else:
            self.global_settings = settings
        self.save_presets()
    
    def get_global_settings(self):
        return self.global_settings
    
    def add_preset(self, name, config):
        self.presets[name] = config
        return self.save_presets()
    
    def remove_preset(self, name):
        if name in self.presets:
            del self.presets[name]
            return self.save_presets()
        return False
    
    def get_preset(self, name):
        return self.presets.get(name, None)
    
    def get_all_presets(self):
        return list(self.presets.keys())


class MainWindow(QMainWindow):
    # Excel导入完成信号: (成功标志, 消息, 职业数量, 是否手动导入)
    signal_excel_loaded = QtCore.pyqtSignal(bool, str, int, bool)

    def __init__(self, start_minimized=False):
        super().__init__()
        self.setWindowTitle("DNF Buff动画随机切换工具")
        self.setGeometry(100, 100, 900, 650)
        
        self.vocation_mapping = VocationMapping()
        self.preset_manager = PresetManager()
        self.switchers = {}
        self.global_filepath = ''
        self.backup_filepath = ''
        self.thumb_dir = os.path.join(APP_DIR, 'thumbnails')
        self.file_manager = GlobalFileManager()
        self._closing = False
        self._loading = False  # 加载预设时禁止触发保存
        
        self.init_ui()
        self.init_tray()

        # 连接Excel导入完成信号（线程安全）
        self.signal_excel_loaded.connect(self._on_excel_loaded)

        # 先恢复全局设置（不依赖预设）
        self._loading = True
        gs = self.preset_manager.get_global_settings()
        self.auto_start_check.setChecked(gs.get('auto_start', False))
        self.boot_start_check.setChecked(gs.get('boot_start', False))
        self.hide_after_start_check.setChecked(gs.get('hide_after_start', False))
        self.close_to_tray_check.setChecked(gs.get('close_to_tray', True))
        if gs.get('filepath', ''):
            self.global_filepath = gs.get('filepath', '')
            self.path_edit.setText(self.global_filepath)
        if gs.get('backup_filepath', ''):
            self.backup_filepath = gs.get('backup_filepath', '')
            self.backup_path_edit.setText(self.backup_filepath)
        self.default_interval_spin.setValue(gs.get('default_interval', 30))

        # 自动导入Excel对应表（后台线程，避免卡死）
        self.auto_import_excel(gs)

        self._loading = False
        
        # 启动时自动加载上次使用的预设（静默）
        if self.preset_manager.last_preset and self.preset_manager.last_preset in self.preset_manager.get_all_presets():
            self.load_preset_by_name(self.preset_manager.last_preset, silent=True)
        
        # 如果配置中启用了开机自启，确保注册表正确
        if gs.get('boot_start', False):
            self.verify_and_fix_boot_start()
        
        # 仅开机自启时最小化到托盘
        if start_minimized:
            self.hide()
    
    def init_tray(self):
        # 用系统标准图标确保托盘可见
        style = self.style()
        icon = style.standardIcon(QtWidgets.QStyle.SP_MediaPlay)
        
        self.tray_icon = QSystemTrayIcon(icon, self)
        self.tray_icon.setToolTip("DNF Buff切换工具")
        
        tray_menu = QMenu()
        show_action = QAction("显示窗口", self)
        show_action.triggered.connect(self.show_from_tray)
        tray_menu.addAction(show_action)
        
        tray_menu.addSeparator()
        
        exit_action = QAction("退出", self)
        exit_action.triggered.connect(self.quit_app)
        tray_menu.addAction(exit_action)
        
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.activated.connect(self.on_tray_activated)
        self.tray_icon.show()
    
    def on_tray_activated(self, reason):
        if reason == QSystemTrayIcon.Trigger:
            self.show_from_tray()
    
    def show_from_tray(self):
        self.show()
        self.activateWindow()
        self.raise_()
    
    def quit_app(self):
        self._closing = True
        self.stop_all()
        self.tray_icon.hide()
        QApplication.quit()
    
    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # 全局设置
        path_group = QGroupBox('全局设置')
        path_layout = QHBoxLayout(path_group)
        
        path_layout.addWidget(QLabel('Buff动画路径:'))
        self.path_edit = QLineEdit()
        path_layout.addWidget(self.path_edit)
        self.path_btn = QPushButton('浏览')
        self.path_btn.clicked.connect(self.browse_global_path)
        path_layout.addWidget(self.path_btn)
        
        path_layout.addWidget(QLabel('默认间隔:'))
        self.default_interval_spin = QSpinBox()
        self.default_interval_spin.setRange(2, 300)
        self.default_interval_spin.setValue(30)
        path_layout.addWidget(self.default_interval_spin)
        
        self.auto_start_check = QCheckBox('自动全部启动')
        self.auto_start_check.stateChanged.connect(self.on_global_setting_changed)
        path_layout.addWidget(self.auto_start_check)
        
        self.boot_start_check = QCheckBox('开机自启动')
        self.boot_start_check.stateChanged.connect(self.on_boot_start_changed)
        path_layout.addWidget(self.boot_start_check)
        
        self.hide_after_start_check = QCheckBox('启动后后台运行')
        self.hide_after_start_check.stateChanged.connect(self.on_global_setting_changed)
        path_layout.addWidget(self.hide_after_start_check)
        
        self.close_to_tray_check = QCheckBox('关闭时后台运行')
        self.close_to_tray_check.setChecked(True)
        self.close_to_tray_check.stateChanged.connect(self.on_global_setting_changed)
        path_layout.addWidget(self.close_to_tray_check)
        
        layout.addWidget(path_group)

        # 备用动画路径
        backup_group = QGroupBox('备用动画库')
        backup_layout = QHBoxLayout(backup_group)
        backup_layout.addWidget(QLabel('备用路径:'))
        self.backup_path_edit = QLineEdit()
        backup_layout.addWidget(self.backup_path_edit)
        self.backup_browse_btn = QPushButton('浏览')
        self.backup_browse_btn.clicked.connect(self.browse_backup_path)
        backup_layout.addWidget(self.backup_browse_btn)
        self.gen_thumb_btn = QPushButton('生成预览图')
        self.gen_thumb_btn.clicked.connect(self.generate_thumbnails)
        backup_layout.addWidget(self.gen_thumb_btn)
        self.gallery_btn = QPushButton('浏览备用库')
        self.gallery_btn.clicked.connect(self.open_gallery)
        backup_layout.addWidget(self.gallery_btn)
        layout.addWidget(backup_group)

        # 职业配置表
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(['职业', '动画分组', '间隔(秒)', '状态', '删除'])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        layout.addWidget(self.table)
        
        # 操作按钮
        btn_layout = QHBoxLayout()
        
        self.add_btn = QPushButton('添加职业')
        self.add_btn.clicked.connect(self.add_row)
        btn_layout.addWidget(self.add_btn)
        
        self.import_btn = QPushButton('导入对应表')
        self.import_btn.clicked.connect(lambda: self.import_excel())
        btn_layout.addWidget(self.import_btn)
        
        self.start_all_btn = QPushButton('全部启动')
        self.start_all_btn.clicked.connect(self.start_all)
        btn_layout.addWidget(self.start_all_btn)
        
        self.stop_all_btn = QPushButton('全部暂停')
        self.stop_all_btn.clicked.connect(self.stop_all)
        btn_layout.addWidget(self.stop_all_btn)
        
        self.restore_btn = QPushButton('还原默认')
        self.restore_btn.clicked.connect(self.restore_default)
        btn_layout.addWidget(self.restore_btn)
        
        layout.addLayout(btn_layout)
        
        # 预设管理
        preset_layout = QHBoxLayout()
        self.preset_combo = QComboBox()
        self.preset_combo.addItems(self.preset_manager.get_all_presets())
        if self.preset_manager.last_preset:
            self.preset_combo.setCurrentText(self.preset_manager.last_preset)
        preset_layout.addWidget(QLabel('预设:'))
        preset_layout.addWidget(self.preset_combo)
        
        self.save_preset_btn = QPushButton('保存预设')
        self.save_preset_btn.clicked.connect(self.save_preset)
        preset_layout.addWidget(self.save_preset_btn)
        
        self.load_preset_btn = QPushButton('加载预设')
        self.load_preset_btn.clicked.connect(self.load_preset)
        preset_layout.addWidget(self.load_preset_btn)
        
        self.del_preset_btn = QPushButton('删除预设')
        self.del_preset_btn.clicked.connect(self.delete_preset)
        preset_layout.addWidget(self.del_preset_btn)
        
        layout.addLayout(preset_layout)
        
        # 使用说明
        help_group = QGroupBox('使用说明')
        help_layout = QVBoxLayout(help_group)
        self.help_text = QTextEdit()
        self.help_text.setReadOnly(True)
        self.help_text.setMaximumHeight(150)
        self.help_text.setHtml("""
        <strong>Q:如何使用？</strong><br>
        A:导入一起下载的对应表，点击添加选取你想随机的职业，并将你想为这个职业随机的动画文件全部重命名为含有同个动画分组的命名，如将为剑魂的全部批量重命名为：RenameSword，那么设定动画分组为：RenameSword就会自动按设定间隔与方式切换剑魂的buff动画<br><br>
        <strong>Q:可以多个职业的BUFF动画都切换吗？</strong><br>
        A:可以，上方添加另一个职业的即可。<br><br>
        <strong>Q:可以为不同职业设定不同动画吗？</strong><br>
        A:可以，将A职业的动画分组为动画分组A，B职业的动画分组为动画分组B，两个动画分组区分即可，但是动画分组不能包含另一个，如剑魂为RenameSword，剑帝为RenameSwordF就会混到一起，想区分开来可以改剑帝的为RenameFSword<br><br>
        <strong>Q:动画分组可以相同吗？</strong><br>
        A:可以，支持多职业共享同一动画分组，自动处理文件冲突。
        """)
        help_layout.addWidget(self.help_text)
        layout.addWidget(help_group)
        
        # 状态栏（替代弹窗提示）
        self.setStatusBar(QStatusBar())
    
    def show_status(self, msg):
        """状态栏显示消息，避免弹窗阻塞"""
        self.statusBar().showMessage(msg, 3000)
    
    def on_global_setting_changed(self):
        """全局设置变化时自动保存"""
        if self._loading:
            return
        settings = {
            'auto_start': self.auto_start_check.isChecked(),
            'boot_start': self.boot_start_check.isChecked(),
            'hide_after_start': self.hide_after_start_check.isChecked(),
            'close_to_tray': self.close_to_tray_check.isChecked(),
            'filepath': self.global_filepath,
            'backup_filepath': self.backup_filepath,
            'default_interval': self.default_interval_spin.value()
        }
        self.preset_manager.save_global_settings(settings)
    
    def on_boot_start_changed(self, state):
        """开机自启动变化时，先设置注册表再保存全局设置"""
        if self._loading:
            return
        self.toggle_boot_start(state)
        self.on_global_setting_changed()
    
    def browse_global_path(self):
        path = QFileDialog.getExistingDirectory(self, '选择Buff动画文件夹')
        if path:
            # 检查是否误选为程序所在目录
            if os.path.abspath(path) == os.path.abspath(APP_DIR):
                QMessageBox.warning(
                    self, '警告', 
                    f'检测到您选择了程序所在目录:\n{path}\n\n'
                    '请选择游戏Buff动画的实际存放路径（通常在游戏安装目录下）。\n'
                    '如果继续使用此路径，动画文件将被复制到程序目录中。'
                )
            self.path_edit.setText(path)
            self.global_filepath = path
            self.on_global_setting_changed()
            self.show_status(f'路径已设置: {path}')

    def browse_backup_path(self):
        path = QFileDialog.getExistingDirectory(self, '选择备用动画文件夹')
        if path:
            self.backup_path_edit.setText(path)
            self.backup_filepath = path
            self.on_global_setting_changed()
            self.show_status(f'备用路径已设置: {path}')

    def generate_thumbnails(self):
        """为备用路径下所有bk2文件生成缩略图"""
        if not self.backup_filepath or not os.path.isdir(self.backup_filepath):
            QMessageBox.warning(self, '警告', '请先设置备用动画路径!')
            return

        ffmpeg = find_ffmpeg()
        if not ffmpeg:
            QMessageBox.warning(self, '错误', '未找到ffmpeg，无法生成预览图!')
            return

        # 收集所有bk2文件
        bk2_files = [f for f in os.listdir(self.backup_filepath)
                     if f.lower().endswith('.bk2')]
        if not bk2_files:
            QMessageBox.information(self, '提示', '备用路径中没有找到.bk2文件!')
            return

        # 创建缩略图目录
        os.makedirs(self.thumb_dir, exist_ok=True)

        # 进度对话框
        progress = QProgressDialog('正在生成预览图...', '取消', 0, len(bk2_files), self)
        progress.setWindowTitle('生成预览图')
        progress.setWindowModality(QtCore.Qt.WindowModal)

        success = 0
        for i, bk2 in enumerate(bk2_files):
            if progress.wasCanceled():
                break
            progress.setLabelText(f'转换中: {bk2}')
            progress.setValue(i)
            QApplication.processEvents()

            bk2_path = os.path.join(self.backup_filepath, bk2)
            thumb_name = os.path.splitext(bk2)[0] + '.png'
            thumb_path = os.path.join(self.thumb_dir, thumb_name)

            # 强制删除旧缩略图，确保重新生成
            if os.path.exists(thumb_path):
                try:
                    os.remove(thumb_path)
                except Exception:
                    pass

            # 先获取视频时长，取中间帧避免开头白屏
            duration = 0
            probe = [ffmpeg, '-i', bk2_path]
            try:
                pr = subprocess.run(probe, capture_output=True, timeout=10,
                                   creationflags=subprocess.CREATE_NO_WINDOW)
                # 从stderr解析Duration: 00:00:05.xx
                import re
                m = re.search(r'Duration:\s*(\d+):(\d+):(\d+\.\d+)',
                             pr.stderr.decode('utf-8', errors='ignore'))
                if m:
                    duration = int(m.group(1)) * 3600 + int(m.group(2)) * 60 + float(m.group(3))
            except Exception:
                pass
            # 取中间位置，至少0.5秒
            seek_pos = max(0.5, duration / 2) if duration > 1 else 0.5

            cmd = [ffmpeg, '-y', '-ss', f'{seek_pos:.2f}', '-i', bk2_path,
                   '-vframes', '1', '-vf', 'scale=160:-1',
                   '-q:v', '2', thumb_path]
            try:
                result = subprocess.run(cmd, capture_output=True, timeout=15,
                                       creationflags=subprocess.CREATE_NO_WINDOW)
                if result.returncode == 0 and os.path.exists(thumb_path):
                    # 验证缩略图不是全白（检查文件大小）
                    if os.path.getsize(thumb_path) > 500:
                        success += 1
                    else:
                        # 如果中间帧也失败，尝试取25%位置
                        seek_pos2 = max(0.3, duration * 0.25) if duration > 1 else 0.3
                        cmd2 = [ffmpeg, '-y', '-ss', f'{seek_pos2:.2f}', '-i', bk2_path,
                                '-vframes', '1', '-vf', 'scale=160:-1',
                                '-q:v', '2', thumb_path]
                        r2 = subprocess.run(cmd2, capture_output=True, timeout=15,
                                           creationflags=subprocess.CREATE_NO_WINDOW)
                        if r2.returncode == 0 and os.path.exists(thumb_path):
                            success += 1
                else:
                    # 最后回退：用select滤镜取第10帧
                    cmd3 = [ffmpeg, '-y', '-i', bk2_path,
                            '-vf', 'select=eq(n\,10),scale=160:-1',
                            '-vframes', '1', '-q:v', '2', thumb_path]
                    r3 = subprocess.run(cmd3, capture_output=True, timeout=15,
                                       creationflags=subprocess.CREATE_NO_WINDOW)
                    if r3.returncode == 0 and os.path.exists(thumb_path):
                        success += 1
            except Exception:
                pass

        progress.setValue(len(bk2_files))
        self.show_status(f'预览图生成完成: {success}/{len(bk2_files)}')
        QMessageBox.information(self, '完成', f'预览图生成完成!\n成功: {success}/{len(bk2_files)}')

    def open_gallery(self):
        """打开备用库画廊"""
        if not self.backup_filepath or not os.path.isdir(self.backup_filepath):
            QMessageBox.warning(self, '警告', '请先设置备用动画路径!')
            return

        # 获取当前所有行（用于分配）
        rows_info = []
        for row in range(self.table.rowCount()):
            config = self.get_row_config(row)
            if config['vocation']:
                rows_info.append({
                    'row': row,
                    'vocation': config['vocation'],
                    'keyword': config['keyword'],
                    'target_filename': config['target_filename']
                })

        if not rows_info:
            QMessageBox.warning(self, '警告', '请先添加至少一个职业!')
            return

        dialog = Bk2GalleryDialog(
            self.backup_filepath, self.thumb_dir, rows_info, self)
        dialog.exec_()
    
    def import_excel(self, filepath=None):
        if filepath is None:
            filepath, _ = QFileDialog.getOpenFileName(self, '导入Excel文件', '', 'Excel文件 (*.xlsx)')
        if filepath:
            self.show_status('正在导入Excel，请稍候...')
            threading.Thread(
                target=self._import_excel_worker,
                args=(filepath,),
                daemon=True
            ).start()

    def _import_excel_worker(self, filepath):
        """后台线程：执行耗时的Excel加载，完成后发信号回主线程"""
        try:
            success = self.vocation_mapping.load_from_excel(filepath)
            if success:
                count = len(self.vocation_mapping.get_vocations())
                self.signal_excel_loaded.emit(True, filepath, count, True)
            else:
                self.signal_excel_loaded.emit(False, 'Excel格式错误或无有效数据', 0, True)
        except Exception as e:
            self.signal_excel_loaded.emit(False, str(e), 0, True)

    @QtCore.pyqtSlot(bool, str, int, bool)
    def _on_excel_loaded(self, success, message, count, is_manual):
        """主线程：更新UI（线程安全）
        is_manual: True为用户手动导入，False为启动自动加载
        """
        if success:
            self.update_vocation_combos()
            if is_manual:
                # 手动导入：保存路径并弹窗提示
                self.show_status(f'导入成功! 共{count}个职业')
                self.preset_manager.save_global_settings({
                    'excel_path': message
                }, merge=True)
                QMessageBox.information(self, '导入成功', f'共导入{count}个职业')
            else:
                # 自动加载：仅状态栏提示，不弹窗
                if message:
                    self.show_status(f'已加载职业对应表，共{count}个职业')
                else:
                    self.show_status(f'已加载默认对应表，共{count}个职业')
        else:
            self.show_status(f'导入失败: {message}')
            if is_manual:
                # 只有手动导入失败才弹窗
                QMessageBox.warning(self, '导入失败', f'无法导入Excel文件:\n{message}')

    def auto_import_excel(self, gs):
        """自动导入Excel对应表：优先使用用户导入路径，否则使用内嵌默认表"""
        self.show_status('正在加载职业对应表...')
        threading.Thread(
            target=self._auto_import_excel_worker,
            args=(gs,),
            daemon=True
        ).start()

    def _auto_import_excel_worker(self, gs):
        """后台线程：自动加载对应表"""
        # 1. 优先使用用户之前导入的路径
        user_excel_path = gs.get('excel_path', '')
        if user_excel_path and os.path.exists(user_excel_path):
            try:
                if self.vocation_mapping.load_from_excel(user_excel_path):
                    count = len(self.vocation_mapping.get_vocations())
                    self.signal_excel_loaded.emit(True, 'user', count, False)
                    return
            except Exception:
                pass

        # 2. 尝试加载默认Excel（程序目录或打包资源目录）
        if os.path.exists(DEFAULT_EXCEL_PATH):
            try:
                if self.vocation_mapping.load_from_excel(DEFAULT_EXCEL_PATH):
                    count = len(self.vocation_mapping.get_vocations())
                    self.signal_excel_loaded.emit(True, '', count, False)
                    return
            except Exception as e:
                self.signal_excel_loaded.emit(False, str(e), 0, False)
                return

        # 3. 都没找到
        self.signal_excel_loaded.emit(False, '未找到对应表，请点击"导入对应表"按钮手动导入', 0, False)
    
    def update_vocation_combos(self):
        vocations = self.vocation_mapping.get_vocations()
        for row in range(self.table.rowCount()):
            combo = self.table.cellWidget(row, 0)
            if combo:
                current_text = combo.currentText()
                combo.clear()
                combo.addItems(vocations)
                if current_text in vocations:
                    combo.setCurrentText(current_text)
                combo.setEnabled(True)
                combo.update()
                combo.show()
    
    def add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        vocation_combo = QComboBox()
        vocation_combo.addItems(self.vocation_mapping.get_vocations())
        vocation_combo.setMaxVisibleItems(30)
        self.table.setCellWidget(row, 0, vocation_combo)

        keyword_edit = QLineEdit()
        self.table.setCellWidget(row, 1, keyword_edit)

        interval_spin = QSpinBox()
        interval_spin.setRange(2, 300)
        interval_spin.setValue(self.default_interval_spin.value())
        self.table.setCellWidget(row, 2, interval_spin)

        start_btn = QPushButton('启动')
        start_btn.clicked.connect(lambda checked, r=row: self.toggle_switcher(r))
        self.table.setCellWidget(row, 3, start_btn)

        delete_btn = QPushButton('删除')
        delete_btn.clicked.connect(lambda checked, r=row: self.delete_row(r))
        delete_btn.setStyleSheet('color: red;')
        self.table.setCellWidget(row, 4, delete_btn)

    def delete_row(self, row):
        if row >= 0:
            self.stop_switcher(row)
            self.table.removeRow(row)
            # 更新switchers字典中的行号
            new_switchers = {}
            for r, switcher in self.switchers.items():
                if r < row:
                    new_switchers[r] = switcher
                elif r > row:
                    new_switchers[r - 1] = switcher
            self.switchers = new_switchers

    def preview_bk2(self, row):
        """预览该职业的bk2动画文件"""
        config = self.get_row_config(row)
        filepath = config.get('filepath', '')
        target_filename = config.get('target_filename', '')
        keyword = config.get('keyword', '')
        vocation = config.get('vocation', '')

        if not filepath or not os.path.isdir(filepath):
            QMessageBox.warning(self, '警告', '请先设置Buff动画路径!')
            return
        if not target_filename and not keyword:
            QMessageBox.warning(self, '警告', '请先选择职业或输入动画分组!')
            return

        # 优先找目标文件
        bk2_path = None
        if target_filename:
            p = os.path.join(filepath, target_filename)
            if os.path.exists(p):
                bk2_path = p

        # 没有目标文件，找含动画分组的备用文件
        if not bk2_path and keyword:
            for f in os.listdir(filepath):
                if (keyword.lower() in f.lower()
                        and f.lower().endswith('.bk2')):
                    bk2_path = os.path.join(filepath, f)
                    break

        if not bk2_path:
            QMessageBox.warning(self, '警告',
                f'未找到可预览的bk2文件!\n\n'
                f'职业: {vocation}\n'
                f'目标文件: {target_filename}\n'
                f'动画分组: {keyword}\n\n'
                f'请在文件夹中确认是否存在对应的bk2文件。')
            return

        dialog = Bk2PreviewDialog(bk2_path, self)
        dialog.exec_()

    def get_row_config(self, row):
        """获取行配置，只返回可序列化的基本类型"""
        vocation_combo = self.table.cellWidget(row, 0)
        keyword_edit = self.table.cellWidget(row, 1)
        interval_spin = self.table.cellWidget(row, 2)
        
        vocation = vocation_combo.currentText() if vocation_combo else ''
        keyword = keyword_edit.text() if keyword_edit else ''
        interval = interval_spin.value() if interval_spin else 30
        filepath = self.global_filepath
        target_filename = self.vocation_mapping.get_target_filename(vocation)
        
        return {
            'vocation': str(vocation),
            'keyword': str(keyword),
            'interval': int(interval),
            'filepath': str(filepath),
            'target_filename': str(target_filename)
        }
    
    def toggle_switcher(self, row):
        btn = self.table.cellWidget(row, 3)
        if not btn:
            return
        
        if btn.text() == '启动':
            config = self.get_row_config(row)
            if not config['vocation']:
                QMessageBox.warning(self, '警告', '请选择职业!')
                return
            if not config['keyword']:
                QMessageBox.warning(self, '警告', '请输入动画分组!')
                return
            if not config['filepath']:
                QMessageBox.warning(self, '警告', '请设置Buff动画路径!')
                return
            if not config['target_filename']:
                QMessageBox.warning(self, '警告', '未找到职业对应的目标文件名!')
                return
            
            self.file_manager.register(config['vocation'], config)
            switcher = BuffSwitcher(config, self.file_manager)
            switcher.start()
            self.switchers[row] = switcher
            btn.setText('停止')
            btn.setStyleSheet('background-color: #4CAF50; color: white;')
            self.show_status(f'{config["vocation"]} 已启动')
            
        else:
            self.stop_switcher(row)
    
    def stop_switcher(self, row):
        if row in self.switchers:
            config = self.get_row_config(row)
            self.file_manager.unregister(config.get('vocation', ''))
            self.switchers[row].stop()
            del self.switchers[row]
        btn = self.table.cellWidget(row, 3)
        if btn:
            btn.setText('启动')
            btn.setStyleSheet('')
    
    def start_all(self):
        if not self.global_filepath:
            QMessageBox.warning(self, '警告', '请先设置Buff动画路径!')
            return
        
        started = 0
        for row in range(self.table.rowCount()):
            btn = self.table.cellWidget(row, 3)
            if btn and btn.text() == '启动':
                config = self.get_row_config(row)
                if config['vocation'] and config['keyword'] and config['target_filename']:
                    self.file_manager.register(config['vocation'], config)
                    switcher = BuffSwitcher(config, self.file_manager)
                    switcher.start()
                    self.switchers[row] = switcher
                    btn.setText('停止')
                    btn.setStyleSheet('background-color: #4CAF50; color: white;')
                    started += 1
        
        self.show_status(f'已启动 {started} 个职业')
    
    def stop_all(self):
        for row in list(self.switchers.keys()):
            if row in self.switchers:
                config = self.get_row_config(row)
                self.file_manager.unregister(config.get('vocation', ''))
                self.switchers[row].stop()
                del self.switchers[row]
                btn = self.table.cellWidget(row, 3)
                if btn:
                    btn.setText('启动')
                    btn.setStyleSheet('')
        self.show_status('已全部停止')
    
    def restore_default(self):
        """还原默认动画文件：自动识别压缩文件或文件夹恢复"""
        if not self.global_filepath:
            QMessageBox.warning(self, '警告', '请先设置Buff动画路径!')
            return
        
        # 使用Directory模式+非原生对话框，可同时选择文件或文件夹
        fd = QFileDialog(self, '选择备份(zip文件或文件夹)')
        fd.setFileMode(QFileDialog.Directory)
        fd.setOption(QFileDialog.DontUseNativeDialog, True)
        fd.setOption(QFileDialog.ShowDirsOnly, False)
        if not fd.exec_():
            return
        
        selected = fd.selectedFiles()
        if not selected:
            return
        
        src_path = selected[0]
        
        # 自动判断类型
        if os.path.isfile(src_path) and src_path.lower().endswith('.zip'):
            reply = QMessageBox.question(self, '确认还原',
                f'将从压缩文件还原动画到:\n{self.global_filepath}\n\n将覆盖同名文件，是否继续?')
            if reply != QMessageBox.Yes:
                return
            try:
                with zipfile.ZipFile(src_path, 'r') as zf:
                    zf.extractall(self.global_filepath)
                self.show_status('还原成功!')
                QMessageBox.information(self, '成功', '动画文件已还原!')
            except Exception as e:
                QMessageBox.warning(self, '错误', f'还原失败:\n{e}')
        elif os.path.isdir(src_path):
            reply = QMessageBox.question(self, '确认还原',
                f'将从文件夹还原动画到:\n{self.global_filepath}\n\n将覆盖同名文件，是否继续?')
            if reply != QMessageBox.Yes:
                return
            try:
                count = 0
                for item in os.listdir(src_path):
                    src = os.path.join(src_path, item)
                    dst = os.path.join(self.global_filepath, item)
                    if os.path.isfile(src):
                        shutil.copy2(src, dst)
                        count += 1
                    elif os.path.isdir(src):
                        if os.path.exists(dst):
                            shutil.rmtree(dst)
                        shutil.copytree(src, dst)
                        count += 1
                self.show_status(f'还原成功! 共{count}个文件/文件夹')
                QMessageBox.information(self, '成功', f'动画文件已还原! 共{count}个文件/文件夹')
            except Exception as e:
                QMessageBox.warning(self, '错误', f'还原失败:\n{e}')
        else:
            QMessageBox.warning(self, '警告', '请选择zip压缩文件或文件夹!')
    
    def save_preset(self):
        """保存预设 - 使用自定义对话框避免阻塞"""
        # 先弹窗获取名称
        dialog = PresetNameDialog(self, '保存预设', '输入预设名称:')
        if dialog.exec_() != QDialog.Accepted:
            return
        
        name = dialog.get_name()
        if not name:
            return
        
        # 收集配置（确保都是基本类型）
        configs = []
        for row in range(self.table.rowCount()):
            try:
                config = self.get_row_config(row)
                if config['vocation'] and config['keyword']:
                    configs.append(config)
            except Exception as e:
                print(f"读取第{row}行配置失败: {e}")
        
        preset_data = {
            'filepath': str(self.global_filepath),
            'backup_filepath': str(self.backup_filepath),
            'default_interval': int(self.default_interval_spin.value()),
            'auto_start': bool(self.auto_start_check.isChecked()),
            'boot_start': bool(self.boot_start_check.isChecked()),
            'hide_after_start': bool(self.hide_after_start_check.isChecked()),
            'close_to_tray': bool(self.close_to_tray_check.isChecked()),
            'configs': configs
        }
        
        # 保存
        if self.preset_manager.add_preset(name, preset_data):
            self.preset_manager.set_last_preset(name)
            # 更新下拉框
            self.preset_combo.clear()
            self.preset_combo.addItems(self.preset_manager.get_all_presets())
            self.preset_combo.setCurrentText(name)
            self.show_status(f'预设 "{name}" 保存成功!')
        else:
            err = self.preset_manager.last_error or '未知错误'
            QMessageBox.warning(self, '错误', f'保存预设失败!\n错误: {err}\n配置路径: {CONFIG_FILE}')
    
    def load_preset(self):
        preset_name = self.preset_combo.currentText()
        if preset_name:
            self.load_preset_by_name(preset_name, silent=False)
    
    def load_preset_by_name(self, preset_name, silent=False):
        preset_data = self.preset_manager.get_preset(preset_name)
        if not preset_data:
            return
        
        self._loading = True  # 加载时禁止触发保存
        self.stop_all()
        
        # 清空表格
        while self.table.rowCount() > 0:
            self.table.removeRow(0)
        
        # 恢复全局设置
        self.global_filepath = preset_data.get('filepath', '')
        self.path_edit.setText(self.global_filepath)
        self.backup_filepath = preset_data.get('backup_filepath', '')
        self.backup_path_edit.setText(self.backup_filepath)
        self.default_interval_spin.setValue(preset_data.get('default_interval', 30))
        self.auto_start_check.setChecked(preset_data.get('auto_start', False))
        self.boot_start_check.setChecked(preset_data.get('boot_start', False))
        self.hide_after_start_check.setChecked(preset_data.get('hide_after_start', False))
        self.close_to_tray_check.setChecked(preset_data.get('close_to_tray', True))
        self._loading = False  # 恢复后保存一次最新状态
        self.on_global_setting_changed()
        
        self.preset_manager.set_last_preset(preset_name)
        self.preset_combo.setCurrentText(preset_name)
        
        # 恢复职业配置
        configs = preset_data.get('configs', [])
        for config in configs:
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            vocation_combo = QComboBox()
            vocation_combo.addItems(self.vocation_mapping.get_vocations())
            vocation_combo.setMaxVisibleItems(30)
            vocation_combo.setCurrentText(config.get('vocation', ''))
            self.table.setCellWidget(row, 0, vocation_combo)
            
            keyword_edit = QLineEdit(config.get('keyword', ''))
            self.table.setCellWidget(row, 1, keyword_edit)
            
            interval_spin = QSpinBox()
            interval_spin.setRange(2, 300)
            interval_spin.setValue(config.get('interval', 30))
            self.table.setCellWidget(row, 2, interval_spin)
            
            start_btn = QPushButton('启动')
            start_btn.clicked.connect(lambda checked, r=row: self.toggle_switcher(r))
            self.table.setCellWidget(row, 3, start_btn)

            delete_btn = QPushButton('删除')
            delete_btn.clicked.connect(lambda checked, r=row: self.delete_row(r))
            delete_btn.setStyleSheet('color: red;')
            self.table.setCellWidget(row, 4, delete_btn)

        # 自动启动
        if self.auto_start_check.isChecked():
            self.start_all()
        
        if not silent:
            self.show_status(f'预设 "{preset_name}" 加载成功!')
    
    def delete_preset(self):
        preset_name = self.preset_combo.currentText()
        if preset_name:
            reply = QMessageBox.question(self, '确认', f'确定删除预设 "{preset_name}" 吗?')
            if reply == QMessageBox.Yes:
                self.preset_manager.remove_preset(preset_name)
                self.preset_combo.removeItem(self.preset_combo.currentIndex())
                if self.preset_manager.last_preset == preset_name:
                    self.preset_manager.set_last_preset('')
                self.show_status('预设已删除')
    
    def verify_and_fix_boot_start(self):
        """启动时校验注册表开机自启，不存在则自动创建"""
        import winreg
        
        exe_path = sys.executable if hasattr(sys, 'frozen') else os.path.abspath(__file__)
        exe_path = os.path.abspath(exe_path)
        cmd_with_args = f'"{exe_path}" --minimized'
        
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                 r'Software\Microsoft\Windows\CurrentVersion\Run',
                                 0, winreg.KEY_READ)
            try:
                existing_value, _ = winreg.QueryValueEx(key, 'DNFBuffSwitcher')
                winreg.CloseKey(key)
                # 值存在，检查路径是否正确
                if cmd_with_args != existing_value:
                    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                         r'Software\Microsoft\Windows\CurrentVersion\Run',
                                         0, winreg.KEY_SET_VALUE)
                    winreg.SetValueEx(key, 'DNFBuffSwitcher', 0, winreg.REG_SZ, cmd_with_args)
                    winreg.CloseKey(key)
            except FileNotFoundError:
                winreg.CloseKey(key)
                # 值不存在，自动创建
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                     r'Software\Microsoft\Windows\CurrentVersion\Run',
                                     0, winreg.KEY_SET_VALUE)
                winreg.SetValueEx(key, 'DNFBuffSwitcher', 0, winreg.REG_SZ, cmd_with_args)
                winreg.CloseKey(key)
        except Exception as e:
            print(f"校验开机自启失败: {e}")
    
    def verify_boot_start(self):
        """兼容旧调用"""
        self.verify_and_fix_boot_start()
    
    def toggle_boot_start(self, state):
        """用注册表实现开机自启（不需要管理员权限，最可靠）"""
        import winreg
        
        exe_path = sys.executable if hasattr(sys, 'frozen') else os.path.abspath(__file__)
        exe_path = os.path.abspath(exe_path)
        cmd_with_args = f'"{exe_path}" --minimized'
        
        try:
            if state == 2:  # 勾选
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                     r'Software\Microsoft\Windows\CurrentVersion\Run',
                                     0, winreg.KEY_SET_VALUE)
                winreg.SetValueEx(key, 'DNFBuffSwitcher', 0, winreg.REG_SZ, cmd_with_args)
                winreg.CloseKey(key)
                # 验证是否写入成功
                key2 = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                      r'Software\Microsoft\Windows\CurrentVersion\Run',
                                      0, winreg.KEY_READ)
                val, _ = winreg.QueryValueEx(key2, 'DNFBuffSwitcher')
                winreg.CloseKey(key2)
                self.show_status('已设置开机自启动')
                QMessageBox.information(self, '开机自启动', 
                    f'已成功设置开机自启动！\n\n注册表路径: HKCU\\...\\Run\\DNFBuffSwitcher\n值: {val}')
            else:  # 取消
                try:
                    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                         r'Software\Microsoft\Windows\CurrentVersion\Run',
                                         0, winreg.KEY_SET_VALUE)
                    winreg.DeleteValue(key, 'DNFBuffSwitcher')
                    winreg.CloseKey(key)
                except FileNotFoundError:
                    pass
                self.show_status('已取消开机自启动')
        except Exception as e:
            self.show_status(f'设置失败: {e}')
            QMessageBox.warning(self, '开机自启动失败', 
                f'设置开机自启动失败！\n\n错误: {e}\n\n请尝试右键exe以管理员身份运行后再设置。')
    
    def closeEvent(self, event):
        if self._closing:
            # 真正退出
            event.accept()
            return
        
        if self.close_to_tray_check.isChecked():
            # 关闭时后台运行：隐藏到托盘
            self.hide()
            event.ignore()
        else:
            # 直接关闭
            self.stop_all()
            self.tray_icon.hide()
            event.accept()


if __name__ == '__main__':
    import socket

    # 单实例检测
    def is_single_instance():
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.bind(('127.0.0.1', 19876))
            return sock
        except OSError:
            return None

    sock = is_single_instance()
    if sock is None:
        QMessageBox.warning(None, '提示', '程序已在运行中！')
        sys.exit(0)
    
    start_minimized = '--minimized' in sys.argv
    
    app = QApplication(sys.argv)
    app.setQuitOnLastWindowClosed(False)
    
    window = MainWindow(start_minimized=start_minimized)
    
    if not start_minimized:
        window.show()
    
    exit_code = app.exec_()
    
    # 关闭时清理子进程
    import os
    import signal
    try:
        parent_pid = os.getpid()
        result = os.popen(f'wmic process where "ParentProcessId={parent_pid}" get ProcessId').read()
        for line in result.strip().split('\n')[1:]:
            line = line.strip()
            if line.isdigit():
                try:
                    os.kill(int(line), signal.SIGTERM)
                except:
                    pass
    except:
        pass
    
    sock.close()
    sys.exit(exit_code)
