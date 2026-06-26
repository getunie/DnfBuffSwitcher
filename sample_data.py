from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "职业对应表"

ws['A1'] = '职业'
ws['B1'] = '目标文件名'

data = [
    ('剑魂', '06_swordman_M_buf_battle'),
    ('剑帝', '06_swordman_F_buf_flow'),
    ('红眼', '06_swordman_M_buf_blood'),
    ('漫游', '06_gunner_M_buf_spit'),
    ('弹药', '06_gunner_M_buf_mine'),
    ('机械', '06_gunner_M_buf_robot'),
    ('大枪', '06_gunner_M_buf_fire'),
    ('剑宗', '06_swordman_F_buf_blade'),
    ('剑魔', '06_swordman_F_buf_demon'),
    ('暗帝', '06_swordman_F_buf_dark'),
    ('剑豪', '06_swordman_F_buf_blossom'),
    ('鬼泣', '06_swordman_M_buf_death'),
    ('阿修罗', '06_swordman_M_buf_eye'),
    ('武神', '06_fighter_F_buf_tiger'),
    ('柔道', '06_fighter_F_buf_grapple'),
    ('百花', '06_fighter_F_buf_zen'),
    ('毒王', '06_fighter_F_buf_poison'),
    ('气功', '06_fighter_M_buf_qi'),
    ('散打', '06_fighter_M_buf_strike'),
    ('街霸', '06_fighter_M_buf_rage'),
    ('元素', '06_mage_F_buf_element'),
    ('召唤', '06_mage_F_buf_summon'),
    ('魔道', '06_mage_F_buf_witch'),
    ('战法', '06_mage_F_buf_spear'),
    ('冰结', '06_mage_M_buf_ice'),
    ('魔皇', '06_mage_M_buf_magic'),
    ('奶爸', '06_priest_M_buf_buff'),
    ('奶妈', '06_priest_F_buf_holy'),
    ('复仇者', '06_priest_M_buf_demon'),
    ('四姨', '06_priest_F_buf_fallen'),
    ('蓝拳', '06_priest_M_buf_fist'),
    ('驱魔', '06_priest_M_buf_exorcist'),
    ('刺客', '06_thief_F_buf_shadow'),
    ('死灵', '06_thief_F_buf_skeleton'),
    ('忍者', '06_thief_F_buf_ninja'),
    ('影舞者', '06_thief_F_buf_dance'),
    ('暗刃', '06_thief_M_buf_blade'),
    ('特工', '06_thief_M_buf_agent'),
    ('战线佣兵', '06_thief_M_buf_war'),
    ('源能专家', '06_thief_M_buf_source'),
    ('征战者', '06_swordman_M_buf_war'),
    ('决战者', '06_swordman_M_buf_spear'),
    ('帕拉丁', '06_swordman_F_buf_paladin'),
    ('龙骑士', '06_swordman_F_buf_dragon'),
    ('狩猎者', '06_gunner_M_buf_hunter'),
    ('暗枪', '06_gunner_M_buf_dark'),
    ('光枪', '06_gunner_M_buf_light'),
    ('合金战士', '06_gunner_F_buf_mech'),
    ('旅人', '06_gunner_F_buf_traveler'),
    ('刃影', '06_thief_F_buf_blade'),
    ('混沌魔灵', '06_mage_F_buf_chaos'),
    ('缔造者', '06_mage_F_buf_creator'),
]

for i, (vocation, filename) in enumerate(data, start=2):
    ws[f'A{i}'] = vocation
    ws[f'B{i}'] = filename

wb.save('职业对应表.xlsx')
print("示例Excel文件已创建: 职业对应表.xlsx")